import openpyxl
from aiogram import types
from config import FORA
from dbase import Database
import pandas as pd
from config import superadmin,ADMIN
from aiogram.types import FSInputFile
from confvar import DBASE,PATH1,PATH2,PATH3

async def report1(message: types.Message,current_date,rp):
    x = f"{rp}-отчет по приходу"
    if (message.from_user.id == superadmin or message.from_user.id==ADMIN ):  # Отправим ADMIN у
        await message.answer(x)
        print(f"{x}. на {current_date}")
        dat = []
        hap = ['Офис','Фио','Должность','Кун','Вакт','Кечикиш минут']
        dat.append(hap)
        path = f"{PATH2}{rp}_{current_date}.xlsx"
        #print(f"{path=}")
        query = """
                SELECT o.name AS office, \
                       u.name AS fio, \
                       u.dolj AS dol, \
                       d.data, \
                       d.time, \
                       d.raz
                FROM davomad d
                         LEFT JOIN users u ON d.user_id = u.user_id
                         LEFT JOIN offices o ON d.office_id = o.id
                WHERE d.data = ? AND d.prz = ?\
                """
        # Выполнение
        async with Database(DBASE) as dbs:
            results = await dbs.fetch_all(query, (current_date,0,))  # Келиш
            for row in results:
                row_list = list(row)
                minit = int(row_list[5])
                if minit < FORA:
                    row_list[5] = ''
                dat.append(row_list)
                #print(row_list)
        async with Database(DBASE) as dbs:
            results = await dbs.fetch_all(query, (current_date,1,))  # Кетишш
            for row in results:
                row_list = list(row)
                minit = int(row_list[5])
                if minit > 0 :
                    row_list[5] = ''
                dat.append(row_list)
        #   =======================================================
        with pd.ExcelWriter(path) as writer:
            # Записываем заголовок в ячейку A1
            txt = f"{current_date} учун давомад"
            pd.DataFrame([txt]).to_excel(writer, index=False, header=False)

            # Создаем DataFrame (первый список станет заголовком, если указать columns)
            df = pd.DataFrame(dat[1:], columns=dat[0])

            # Записываем таблицу, начиная со второй строки (индекс 1)
            df.to_excel(writer, startrow=1, index=False)

        file = FSInputFile(path)             # Загружаем файл
        await message.answer_document(file)  # Отправляем файл
    else:
        ss = f"{x}-задана команда ADMIN"
        print(ss)

#-----------------------------------------------------------------------------------------------
async def spis(message: types.Message,current_date):
    # Загружаем файл
    path = PATH1 + 'coworkers.xlsx'
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet = wb.active

    data = [list(row) for row in sheet.iter_rows(values_only=True)]  # Превращаем строки в список списков

    path1 = f"{PATH2}spisok.txt"
    with open(path1, "w", encoding="utf-8") as file:
        for i, lst in enumerate(data):
            if i > 0:
                ofn = "офис не найден"
                x = "Не регистрирован"
                async with Database(DBASE) as dbs:
                    st = "SELECT * FROM users WHERE phone = ?"
                    user = await dbs.fetch_one(st, (lst[1],))  # из users
                    st = "SELECT * FROM offices WHERE id = ?"
                    ofs = await dbs.fetch_one(st, (lst[4],))  # из users
                    if ofs:
                        ofn = ofs[1]

                    if user:
                        #print(f"Найден {s} {ofn}")
                        x = "Регистрирован   "
                ss = x +  f" {lst[1]:>14} {lst[2]:<35} {lst[3]:<15} {ofn:<20}\n"
                file.write(ss)
                #print(f"{ss}")
    file = FSInputFile(path1)             # Загружаем файл
    await message.answer_document(file)  # Отправляем файл

#--------------------------Келди кетди на datdav------------------------
async def keldi_ketdi(message: types.Message,datdav):
    s = f"{datdav} га давомад"
    print(s)
    await message.answer(s)
    #------------------------------------------------
    davm = []
    # Загружаем шапку
    davm.append(f"{datdav} учун келди-кетди руйхати")
    txt = []
    # Загружаем файл
    path = PATH1 + 'coworkers.xlsx'
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet = wb.active

    #-------------------------------------------------
    data = [list(row) for row in sheet.iter_rows(values_only=True)]  # Превращаем строки в список списков
    data[1:] = sorted(data[1:], key=lambda x: x[4])
    #-------------------------------------------------
    path3 = PATH3 + 'keldi_ketdi.txt'
    print("Открываем", path3)
    with open(path3, 'r', encoding='utf-8') as file:
        for line in file:
            txt.append(line)
            tit = line.rstrip()
            davm.append(tit)
            #print(tit)
    for i, lst in enumerate(data):
        if i > 0:  # Шапка не считается
            #lst[1] # 123456789_1234 tel
            #lst[2] # 123456789_123456789_123456789_123456789_ name
            #lst[3] # 123456789_123456789_dolj
            #lst[4] #  office_id
            #print(lst)
            n = str(i)
            #========================================================
            async with Database(DBASE) as dbs:
                st = "SELECT * FROM users WHERE phone = ?"
                user = await dbs.fetch_one(st, (lst[1],))  # из users
                st = "SELECT * FROM offices WHERE id = ?"
                ofs = await dbs.fetch_one(st, (lst[4],))  # из users
                ofn = 'Нет такой офис'
                if ofs:
                    ofn = ofs[1]
                dt1 = '          '
                tm1 = '     '
                rz1 = '    '
                dt2 = '          '
                tm2 = '     '
                rz2 = '    '
                if user:
                    x = ""
                    # Из davomad по user[0]=users->user_id   davomad->user_id==user[0]
                    #                                        davomad->data == datdav
                    #
#-----------------------------------------------------------------------------------------------
                    query = """SELECT data, time,raz FROM davomad WHERE user_id = ? AND data = ? AND prz = ?"""
                    tup = (user[0],datdav,0)    # Келиш
                    rez = await dbs.fetch_one(query, tup )  # из davomad приход
                    if rez:
                        dt1 = rez[0]
                        tm1 = rez[1][:5]
                        zx = int(rez[2])
                        if zx>0:
                            rz1 = str(rez[2])

                        tp1 = (user[0], datdav, 1) # Кетиш

                        res = await dbs.fetch_one(query, tp1)  # из davomad уход
                        if res:
                            dt2 = res[0]
                            tm2 = res[1][:5]
                            zx = int(res[2])
                            if zx < 0:
                                rz2 = str(res[2])
                else:
                    x = 'Не регистрирован'
            #========================================================
            ss = f"{n:>3} {lst[1]:<14} {lst[2]:<40} {lst[3]:<20} {tm1} {rz1:>8}   {tm2} {rz2:>8}   {ofn:<20} {x}"
            davm.append(ss)
    #-----------------------------------------------------------------------
    path2 = f"{PATH2}{datdav}_davomad.txt"
    with open(path2, "w", encoding="utf-8") as file:
        for lin in  davm:
            sf = f"{lin}\n"
            file.write(sf)
    file = FSInputFile(path2)             # Загружаем файл
    await message.answer_document(file)  # Отправляем файл

